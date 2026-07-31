#!/usr/bin/env python3
"""Gate 4 private-input contracts and privacy-safe validation.

Raw portfolio values remain in memory. Validation results identify fields and
checks without copying policy values, holdings, or opportunity-set rows into
diagnostic output.
"""

from __future__ import annotations

import csv
import json
import math
from dataclasses import asdict, dataclass
from datetime import UTC, date, datetime
from pathlib import Path
from typing import Any, Callable

try:
    import yaml
except ImportError:  # pragma: no cover - exercised through dependency diagnostics
    yaml = None

try:
    from jsonschema import Draft202012Validator, FormatChecker
except ImportError:  # pragma: no cover - exercised through dependency diagnostics
    Draft202012Validator = None
    FormatChecker = None

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - exercised through dependency diagnostics
    load_workbook = None


SCRIPT_DIR = Path(__file__).resolve().parent
GATE4_DIR = SCRIPT_DIR.parent / "gate4"
SCHEMA_DIR = GATE4_DIR / "schemas"
FIELD_GOVERNANCE_PATH = GATE4_DIR / "field_governance.json"
PRIVATE_INPUT_CONTRACT_VERSION = "2.1.0"
FRAMEWORK_STATUS = "GATE_4_FRAMEWORK_READY"
INPUT_STATUS_REQUIRED = "GATE_4_PRIVATE_INPUTS_REQUIRED"
INPUT_STATUS_VALIDATED = "GATE_4_INPUTS_VALIDATED"
SUPPORTED_CLASSIFICATIONS = {"PRIVATE_PORTFOLIO", "SYNTHETIC_PUBLIC_EXAMPLE"}
SUPPORTED_INPUT_MODES = {
    "EXPOSURE_ONLY",
    "AGGREGATED_PORTFOLIO",
    "FULL_HOLDINGS",
}
REQUIREMENT_CLASSES = {
    "CORE_REQUIRED",
    "CONDITIONAL",
    "OPTIONAL",
    "REVIEWER_CONFIRMED_NOT_APPLICABLE",
}
DOCUMENT_MANIFEST_FIELDS = {
    "policy": "portfolio_policy",
    "exposures": "exposure_summary",
    "holdings": "current_holdings",
    "opportunities": "opportunity_set",
    "constraints": "portfolio_constraint_inputs",
    "approval": "approval_config",
    "freshness": "gate3_freshness_attestation",
}
MODE_CAPABILITIES = {
    "EXPOSURE_ONLY": {
        "aggregate_exposure_validation": "AVAILABLE",
        "holdings_reconciliation": "NOT_EVALUATED",
        "security_level_liquidity": "NOT_EVALUATED",
    },
    "AGGREGATED_PORTFOLIO": {
        "aggregate_exposure_validation": "AVAILABLE",
        "holdings_reconciliation": "AVAILABLE",
        "security_level_liquidity": "NOT_EVALUATED",
    },
    "FULL_HOLDINGS": {
        "aggregate_exposure_validation": "DERIVABLE_FROM_HOLDINGS",
        "holdings_reconciliation": "AVAILABLE",
        "security_level_liquidity": "AVAILABLE",
    },
}
ALLOWED_DECISIONS = {
    "PENDING",
    "APPROVED",
    "APPROVED_WITH_MODIFICATION",
    "REJECTED",
    "DEFERRED",
}


@dataclass
class PrivateInputCheck:
    check_id: str
    document: str
    status: str
    issue_class: str
    field: str
    row_number: int | None
    message: str
    remediation: str

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


@dataclass
class PrivateInputBundle:
    manifest: dict[str, Any]
    policy: dict[str, Any]
    exposures: list[dict[str, Any]]
    holdings: list[dict[str, Any]]
    opportunities: list[dict[str, Any]]
    constraint_inputs: dict[str, Any]
    approval: dict[str, Any]
    freshness_attestation: dict[str, Any]
    field_governance_contract: dict[str, Any]
    manifest_path: Path
    output_dir: Path


class PrivateInputLoadError(ValueError):
    """Raised with a sanitized message that contains no private values."""


if yaml is not None:
    class UniqueKeySafeLoader(yaml.SafeLoader):
        pass


    def _construct_unique_mapping(loader: Any, node: Any, deep: bool = False) -> dict[Any, Any]:
        mapping: dict[Any, Any] = {}
        for key_node, value_node in node.value:
            key = loader.construct_object(key_node, deep=deep)
            if key in mapping:
                raise PrivateInputLoadError("The YAML mapping contains a duplicate key.")
            mapping[key] = loader.construct_object(value_node, deep=deep)
        return mapping


    UniqueKeySafeLoader.add_constructor(
        yaml.resolver.BaseResolver.DEFAULT_MAPPING_TAG,
        _construct_unique_mapping,
    )


def utc_now() -> str:
    return datetime.now(UTC).strftime("%Y-%m-%dT%H:%M:%SZ")


def parse_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    try:
        return date.fromisoformat(str(value)[:10])
    except ValueError:
        return None


def parse_datetime(value: Any) -> datetime | None:
    if value is None or value == "":
        return None
    text = str(value).strip().replace("Z", "+00:00")
    try:
        parsed = datetime.fromisoformat(text)
    except ValueError:
        return None
    if parsed.tzinfo is None:
        return parsed.replace(tzinfo=UTC)
    return parsed.astimezone(UTC)


def parse_number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        parsed = float(value)
        return parsed if math.isfinite(parsed) else None
    text = str(value).strip().replace(",", "")
    try:
        parsed = float(text)
    except ValueError:
        return None
    return parsed if math.isfinite(parsed) else None


def parse_ratio(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, str) and value.strip().endswith("%"):
        parsed = parse_number(value.strip()[:-1])
        return parsed / 100 if parsed is not None else None
    return parse_number(value)


def parse_integer(value: Any) -> int | None:
    parsed = parse_number(value)
    if parsed is None or not float(parsed).is_integer():
        return None
    return int(parsed)


def clean_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def read_mapping(path: Path) -> dict[str, Any]:
    suffix = path.suffix.lower()
    try:
        text = path.read_text(encoding="utf-8")
    except OSError as exc:
        raise PrivateInputLoadError("The mapping file could not be read.") from exc
    try:
        if suffix == ".json":
            def unique_pairs(pairs: list[tuple[str, Any]]) -> dict[str, Any]:
                mapping: dict[str, Any] = {}
                for key, value in pairs:
                    if key in mapping:
                        raise PrivateInputLoadError(
                            "The JSON object contains a duplicate key."
                        )
                    mapping[key] = value
                return mapping

            payload = json.loads(text, object_pairs_hook=unique_pairs)
        elif suffix in {".yaml", ".yml"}:
            if yaml is None:
                raise PrivateInputLoadError(
                    "YAML support is unavailable; install requirements-gate4.txt."
                )
            payload = yaml.load(text, Loader=UniqueKeySafeLoader)
        else:
            raise PrivateInputLoadError("Only JSON and YAML mapping files are supported.")
    except PrivateInputLoadError:
        raise
    except (json.JSONDecodeError, getattr(yaml, "YAMLError", ValueError)) as exc:
        raise PrivateInputLoadError("The mapping file is not valid JSON or YAML.") from exc
    if not isinstance(payload, dict):
        raise PrivateInputLoadError("The mapping file must contain one object.")
    return payload


def _csv_rows(path: Path) -> list[dict[str, Any]]:
    try:
        with path.open(newline="", encoding="utf-8-sig") as handle:
            reader = csv.DictReader(handle)
            if not reader.fieldnames:
                raise PrivateInputLoadError("The table has no header row.")
            headers = [str(value).strip() for value in reader.fieldnames]
            if any(not header for header in headers) or len(headers) != len(set(headers)):
                raise PrivateInputLoadError(
                    "Every CSV header must be populated and unique."
                )
            rows: list[dict[str, Any]] = []
            for row in reader:
                if None in row and any(
                    clean_text(value) is not None for value in (row.get(None) or [])
                ):
                    raise PrivateInputLoadError(
                        "A CSV row contains values beyond the declared header."
                    )
                cleaned = {
                    str(key).strip(): clean_text(value)
                    for key, value in row.items()
                    if key is not None
                }
                if any(value is not None for value in cleaned.values()):
                    rows.append(cleaned)
            return rows
    except UnicodeDecodeError as exc:
        raise PrivateInputLoadError("The CSV file must use UTF-8 encoding.") from exc
    except OSError as exc:
        raise PrivateInputLoadError("The CSV file could not be read.") from exc


def _xlsx_rows(path: Path) -> list[dict[str, Any]]:
    if load_workbook is None:
        raise PrivateInputLoadError(
            "Excel support is unavailable; install requirements-gate4.txt."
        )
    try:
        workbook = load_workbook(path, read_only=True, data_only=False)
    except (OSError, ValueError) as exc:
        raise PrivateInputLoadError("The Excel workbook could not be read.") from exc
    try:
        worksheet = workbook.active
        raw_rows = list(worksheet.iter_rows())
        if not raw_rows:
            raise PrivateInputLoadError("The table has no header row.")
        if any(
            getattr(cell, "data_type", None) == "f"
            for cells in raw_rows
            for cell in cells
        ):
            raise PrivateInputLoadError(
                "Spreadsheet formulas are prohibited in Gate 4 input tables."
            )
        headers = [clean_text(cell.value) for cell in raw_rows[0]]
        if (
            not headers
            or any(header is None for header in headers)
            or len(headers) != len(set(headers))
        ):
            raise PrivateInputLoadError(
                "Every Excel header cell must be populated and unique."
            )
        rows: list[dict[str, Any]] = []
        for cells in raw_rows[1:]:
            values = [cell.value for cell in cells]
            if not any(clean_text(value) is not None for value in values):
                continue
            rows.append(
                {
                    str(header): clean_text(value) if isinstance(value, str) else value
                    for header, value in zip(headers, values)
                }
            )
        return rows
    finally:
        workbook.close()


def read_table(path: Path) -> list[dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _csv_rows(path)
    if suffix == ".xlsx":
        return _xlsx_rows(path)
    raise PrivateInputLoadError("Only CSV and XLSX table files are supported.")


def normalize_holding_row(row: dict[str, Any]) -> dict[str, Any]:
    normalized = dict(row)
    for field in (
        "position_weight",
        "market_value_base_currency",
        "average_daily_value_traded",
        "estimated_days_to_exit",
        "hedge_ratio",
    ):
        parser = parse_ratio if field in {"position_weight", "hedge_ratio"} else parse_number
        normalized[field] = parser(row.get(field))
    normalized["existing_hedge_identifier"] = clean_text(row.get("existing_hedge_identifier"))
    normalized["security_identifier"] = clean_text(row.get("security_identifier"))
    return normalized


def normalize_exposure_row(row: dict[str, Any]) -> dict[str, Any]:
    normalized = dict(row)
    normalized["exposure_weight"] = parse_ratio(row.get("exposure_weight"))
    normalized["market_value_base_currency"] = parse_number(
        row.get("market_value_base_currency")
    )
    normalized["source_locator"] = clean_text(row.get("source_locator"))
    return normalized


def normalize_opportunity_row(row: dict[str, Any]) -> dict[str, Any]:
    normalized = dict(row)
    for field in ("expected_return", "bear_case_downside"):
        normalized[field] = parse_ratio(row.get(field))
    normalized["holding_period_days"] = parse_integer(row.get("holding_period_days"))
    normalized["estimated_days_to_exit"] = parse_number(row.get("estimated_days_to_exit"))
    normalized["source_contract_hash"] = clean_text(row.get("source_contract_hash"))
    return normalized


def load_schema(name: str) -> dict[str, Any]:
    path = SCHEMA_DIR / name
    return json.loads(path.read_text(encoding="utf-8"))


def load_field_governance_contract() -> dict[str, Any]:
    try:
        payload = json.loads(FIELD_GOVERNANCE_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise PrivateInputLoadError(
            "The public field-governance contract could not be loaded."
        ) from exc
    if not isinstance(payload, dict):
        raise PrivateInputLoadError(
            "The public field-governance contract must contain one object."
        )
    return payload


def _field_path(error: Any) -> str:
    parts = [str(value) for value in error.absolute_path]
    if error.validator == "required" and isinstance(error.validator_value, list):
        missing = [
            str(field)
            for field in error.validator_value
            if not isinstance(error.instance, dict) or field not in error.instance
        ]
        if missing:
            parts.append(",".join(sorted(missing)))
    return ".".join(parts) or "<document>"


def schema_checks(
    document: str,
    instance: Any,
    schema_name: str,
    *,
    row_number: int | None = None,
) -> list[PrivateInputCheck]:
    if Draft202012Validator is None or FormatChecker is None:
        return [
            PrivateInputCheck(
                check_id=f"G4I-{document}-dependency",
                document=document,
                status="FAIL",
                issue_class="HARD_STOP",
                field="<dependency>",
                row_number=row_number,
                message="JSON Schema validation is unavailable.",
                remediation="Install requirements-gate4.txt before validating private inputs.",
            )
        ]
    schema = load_schema(schema_name)
    validator = Draft202012Validator(schema, format_checker=FormatChecker())
    errors = sorted(
        validator.iter_errors(instance),
        key=lambda error: (
            tuple(str(value) for value in error.absolute_path),
            str(error.validator),
        ),
    )
    if not errors:
        return [
            PrivateInputCheck(
                check_id=f"G4I-{document}-schema",
                document=document,
                status="PASS",
                issue_class="INFO",
                field="<document>",
                row_number=row_number,
                message="Schema validation passed.",
                remediation="None.",
            )
        ]
    checks: list[PrivateInputCheck] = []
    for index, error in enumerate(errors, start=1):
        field = _field_path(error)
        checks.append(
            PrivateInputCheck(
                check_id=f"G4I-{document}-schema-{row_number or 0}-{index}",
                document=document,
                status="FAIL",
                issue_class="HARD_STOP",
                field=field,
                row_number=row_number,
                message=f"Schema rule `{error.validator}` failed for `{field}`.",
                remediation="Correct the named field using the public field-definition guide.",
            )
        )
    return checks


def add_check(
    checks: list[PrivateInputCheck],
    *,
    check_id: str,
    document: str,
    passed: bool,
    field: str,
    message_pass: str,
    message_fail: str,
    remediation: str,
    row_number: int | None = None,
    issue_class: str = "HARD_STOP",
) -> None:
    checks.append(
        PrivateInputCheck(
            check_id=check_id,
            document=document,
            status="PASS" if passed else "FAIL",
            issue_class="INFO" if passed else issue_class,
            field=field,
            row_number=row_number,
            message=message_pass if passed else message_fail,
            remediation="None." if passed else remediation,
        )
    )


def _is_within(path: Path, parent: Path) -> bool:
    try:
        path.relative_to(parent)
        return True
    except ValueError:
        return False


def resolve_manifest_path(base: Path, relative_value: Any, *, output: bool = False) -> Path:
    text = clean_text(relative_value)
    if not text:
        raise PrivateInputLoadError("A required manifest file path is missing.")
    relative = Path(text)
    if relative.is_absolute() or ".." in relative.parts:
        raise PrivateInputLoadError("Manifest paths must be relative and cannot traverse directories.")
    resolved = (base / relative).resolve()
    if not _is_within(resolved, base.resolve()):
        raise PrivateInputLoadError("Manifest paths must remain inside the selected private workspace.")
    if not output and not resolved.is_file():
        raise PrivateInputLoadError("A required private input file is missing.")
    return resolved


def _load_bundle_documents(
    manifest_path: Path,
    manifest: dict[str, Any],
    field_governance_contract: dict[str, Any],
    checks: list[PrivateInputCheck],
) -> PrivateInputBundle | None:
    base = manifest_path.resolve().parent
    files = manifest.get("files")
    if not isinstance(files, dict):
        return None
    mode = manifest.get("input_mode")
    mode_requirements = field_governance_contract.get(
        "mode_document_requirements", {}
    ).get(mode, {})
    required_documents = set(mode_requirements.get("required_documents", []))
    optional_documents = set(mode_requirements.get("optional_documents", []))
    prohibited_documents = set(mode_requirements.get("prohibited_documents", []))
    loaders: dict[str, Callable[[Path], Any]] = {
        "policy": read_mapping,
        "exposures": read_table,
        "holdings": read_table,
        "opportunities": read_table,
        "constraints": read_mapping,
        "approval": read_mapping,
        "freshness": read_mapping,
    }
    loaded: dict[str, Any] = {}
    for document, manifest_field in DOCUMENT_MANIFEST_FIELDS.items():
        supplied_path = files.get(manifest_field)
        if document in prohibited_documents:
            if clean_text(supplied_path) is not None:
                checks.append(
                    PrivateInputCheck(
                        check_id=f"G4I-{document}-prohibited-for-mode",
                        document="manifest",
                        status="FAIL",
                        issue_class="HARD_STOP",
                        field=f"files.{manifest_field}",
                        row_number=None,
                        message="The document is prohibited for the selected input mode.",
                        remediation="Remove the prohibited file reference or select the correct input mode.",
                    )
                )
            continue
        if document not in required_documents and document not in optional_documents:
            checks.append(
                PrivateInputCheck(
                    check_id=f"G4I-{document}-undefined-for-mode",
                    document="manifest",
                    status="FAIL",
                    issue_class="HARD_STOP",
                    field=f"files.{manifest_field}",
                    row_number=None,
                    message="The document has no governance rule for the selected input mode.",
                    remediation="Use a supported input mode and the published field-governance contract.",
                )
            )
            continue
        if document in optional_documents and clean_text(supplied_path) is None:
            continue
        try:
            path = resolve_manifest_path(base, supplied_path)
            loaded[document] = loaders[document](path)
        except PrivateInputLoadError as exc:
            checks.append(
                PrivateInputCheck(
                    check_id=f"G4I-{document}-load",
                    document=document,
                    status="FAIL",
                    issue_class="HARD_STOP",
                    field=f"files.{manifest_field}",
                    row_number=None,
                    message=str(exc),
                    remediation="Restore the required local file using the public template.",
                )
            )
    try:
        output_dir = resolve_manifest_path(base, files.get("private_output_dir"), output=True)
    except PrivateInputLoadError as exc:
        checks.append(
            PrivateInputCheck(
                check_id="G4I-private-output-dir",
                document="manifest",
                status="FAIL",
                issue_class="HARD_STOP",
                field="files.private_output_dir",
                row_number=None,
                message=str(exc),
                remediation="Use a relative private output directory inside the private workspace.",
            )
        )
        return None
    if not required_documents.issubset(loaded):
        return None
    return PrivateInputBundle(
        manifest=manifest,
        policy=loaded["policy"],
        exposures=[
            normalize_exposure_row(row)
            for row in loaded.get("exposures", [])
        ],
        holdings=[
            normalize_holding_row(row)
            for row in loaded.get("holdings", [])
        ],
        opportunities=[normalize_opportunity_row(row) for row in loaded["opportunities"]],
        constraint_inputs=loaded["constraints"],
        approval=loaded["approval"],
        freshness_attestation=loaded["freshness"],
        field_governance_contract=field_governance_contract,
        manifest_path=manifest_path.resolve(),
        output_dir=output_dir,
    )


def _governance_definition_checks(
    governance: dict[str, Any],
    checks: list[PrivateInputCheck],
) -> None:
    checks.extend(
        schema_checks(
            "field-governance-contract",
            governance,
            "field_governance.schema.json",
        )
    )
    declared_classes = set(governance.get("requirement_classes", []))
    rules = governance.get("field_rules", [])
    seen_rules: set[tuple[str, str, str]] = set()
    duplicate_rule = False
    if isinstance(rules, list):
        for rule in rules:
            if not isinstance(rule, dict):
                continue
            for mode in rule.get("modes", []):
                key = (
                    str(rule.get("document")),
                    str(rule.get("field_path")),
                    str(mode),
                )
                if key in seen_rules:
                    duplicate_rule = True
                seen_rules.add(key)
    mode_requirements = governance.get("mode_document_requirements", {})
    document_partition_valid = True
    expected_documents = set(DOCUMENT_MANIFEST_FIELDS)
    for mode in SUPPORTED_INPUT_MODES:
        requirement = (
            mode_requirements.get(mode, {})
            if isinstance(mode_requirements, dict)
            else {}
        )
        required = set(requirement.get("required_documents", []))
        optional = set(requirement.get("optional_documents", []))
        prohibited = set(requirement.get("prohibited_documents", []))
        if (
            required.intersection(optional)
            or required.intersection(prohibited)
            or optional.intersection(prohibited)
            or required.union(optional, prohibited) != expected_documents
        ):
            document_partition_valid = False
    add_check(
        checks,
        check_id="G4I-field-governance-definition",
        document="field-governance-contract",
        passed=(
            declared_classes == REQUIREMENT_CLASSES
            and governance.get("schema_required_default_class")
            == "CORE_REQUIRED"
            and not duplicate_rule
            and document_partition_valid
        ),
        field="<contract>",
        message_pass="The shared field-governance contract is complete and internally consistent.",
        message_fail="The shared field-governance contract is incomplete or internally inconsistent.",
        remediation="Correct the shared contract; do not patch a company or private workspace.",
    )


def _nested_value(payload: dict[str, Any], field_path: str) -> Any:
    value: Any = payload
    for part in field_path.split("."):
        if not isinstance(value, dict) or part not in value:
            return None
        value = value[part]
    return value


def _value_is_present(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, str):
        return bool(value.strip())
    return True


def _governed_payloads(
    bundle: PrivateInputBundle,
    document: str,
) -> list[tuple[str | None, int | None, dict[str, Any]]]:
    if document == "manifest":
        return [(None, None, bundle.manifest)]
    if document == "policy":
        return [(None, None, bundle.policy)]
    if document == "approval":
        return [(None, None, bundle.approval)]
    if document == "freshness":
        return [(None, None, bundle.freshness_attestation)]
    if document == "constraints":
        return [(None, None, bundle.constraint_inputs)]
    row_config = {
        "exposures": (bundle.exposures, "exposure_id"),
        "holdings": (bundle.holdings, "position_id"),
        "opportunities": (bundle.opportunities, "opportunity_id"),
    }
    rows, identifier_field = row_config.get(document, ([], ""))
    return [
        (clean_text(row.get(identifier_field)), index, row)
        for index, row in enumerate(rows, start=2)
    ]


def _field_governance_checks(
    bundle: PrivateInputBundle,
    checks: list[PrivateInputCheck],
) -> None:
    mode = bundle.manifest.get("input_mode")
    rules = [
        rule
        for rule in bundle.field_governance_contract.get("field_rules", [])
        if isinstance(rule, dict) and mode in rule.get("modes", [])
    ]
    governance = bundle.manifest.get("field_governance", {})
    exceptions = (
        governance.get("reviewer_confirmed_not_applicable", [])
        if isinstance(governance, dict)
        else []
    )
    exception_map: dict[tuple[str, str, str | None], dict[str, Any]] = {}
    governance_failed = False
    manifest_date = parse_date(bundle.manifest.get("as_of_date"))

    for exception in exceptions:
        if not isinstance(exception, dict):
            continue
        key = (
            str(exception.get("document")),
            str(exception.get("field_path")),
            clean_text(exception.get("row_id")),
        )
        reviewed_at = parse_datetime(exception.get("reviewed_at"))
        matching_rules = [
            rule
            for rule in rules
            if rule.get("document") == key[0]
            and rule.get("field_path") == key[1]
            and rule.get("requirement_class")
            == "REVIEWER_CONFIRMED_NOT_APPLICABLE"
        ]
        payloads = _governed_payloads(bundle, key[0])
        matching_payloads = [
            payload
            for row_id, _, payload in payloads
            if row_id == key[2]
        ]
        valid = (
            key not in exception_map
            and bool(matching_rules)
            and bool(matching_payloads)
            and reviewed_at is not None
            and manifest_date is not None
            and reviewed_at.date() <= manifest_date
        )
        if not valid:
            governance_failed = True
            checks.append(
                PrivateInputCheck(
                    check_id=f"G4I-field-governance-exception-{len(checks) + 1}",
                    document=key[0],
                    status="FAIL",
                    issue_class="HARD_STOP",
                    field=key[1],
                    row_number=None,
                    message="A reviewer-confirmed not-applicable exception is invalid, duplicated, or not permitted.",
                    remediation="Use an allowed field, an existing row, a dated reviewer, and a non-empty rationale.",
                )
            )
        else:
            exception_map[key] = exception

    for rule in rules:
        document = str(rule.get("document"))
        field_path = str(rule.get("field_path"))
        requirement_class = rule.get("requirement_class")
        for row_id, row_number, payload in _governed_payloads(bundle, document):
            value_present = _value_is_present(_nested_value(payload, field_path))
            key = (document, field_path, row_id)
            exception_present = key in exception_map
            if requirement_class == "CORE_REQUIRED" and not value_present:
                governance_failed = True
                checks.append(
                    PrivateInputCheck(
                        check_id=f"G4I-field-core-required-{len(checks) + 1}",
                        document=document,
                        status="FAIL",
                        issue_class="HARD_STOP",
                        field=field_path,
                        row_number=row_number,
                        message="A CORE_REQUIRED field is missing.",
                        remediation="Provide the field locally; reviewer exceptions cannot replace core data.",
                    )
                )
            elif requirement_class == "REVIEWER_CONFIRMED_NOT_APPLICABLE":
                valid = (value_present and not exception_present) or (
                    not value_present and exception_present
                )
                if not valid:
                    governance_failed = True
                    checks.append(
                        PrivateInputCheck(
                            check_id=f"G4I-field-not-applicable-{len(checks) + 1}",
                            document=document,
                            status="FAIL",
                            issue_class="HARD_STOP",
                            field=field_path,
                            row_number=row_number,
                            message="The field is missing without a valid reviewer-confirmed not-applicable record, or a record conflicts with a supplied value.",
                            remediation="Supply the field or add one dated, row-specific reviewer exception, but not both.",
                        )
                    )
    if not governance_failed:
        checks.append(
            PrivateInputCheck(
                check_id="G4I-field-governance",
                document="bundle",
                status="PASS",
                issue_class="INFO",
                field="<governance>",
                row_number=None,
                message="Mode-specific field governance passed.",
                remediation="None.",
            )
        )


def _validate_document_schemas(
    bundle: PrivateInputBundle,
    checks: list[PrivateInputCheck],
) -> None:
    mode = bundle.manifest.get("input_mode")
    mode_requirements = bundle.field_governance_contract.get(
        "mode_document_requirements", {}
    ).get(mode, {})
    required_documents = set(mode_requirements.get("required_documents", []))
    checks.extend(schema_checks("policy", bundle.policy, "portfolio_policy.schema.json"))
    checks.extend(
        schema_checks(
            "constraints",
            bundle.constraint_inputs,
            "portfolio_constraint_inputs.schema.json",
        )
    )
    checks.extend(schema_checks("approval", bundle.approval, "approval_config.schema.json"))
    checks.extend(
        schema_checks(
            "freshness",
            bundle.freshness_attestation,
            "gate3_freshness_attestation.schema.json",
        )
    )
    if "exposures" in required_documents and not bundle.exposures:
        checks.append(
            PrivateInputCheck(
                "G4I-exposures-empty",
                "exposures",
                "FAIL",
                "HARD_STOP",
                "<rows>",
                None,
                "The exposure table contains no exposure rows.",
                "Provide dated aggregate exposures using the public exposure template.",
            )
        )
    for index, row in enumerate(bundle.exposures, start=2):
        checks.extend(
            schema_checks(
                "exposures",
                row,
                "exposure_row.schema.json",
                row_number=index,
            )
        )
    if "holdings" in required_documents and not bundle.holdings:
        checks.append(
            PrivateInputCheck(
                "G4I-holdings-empty",
                "holdings",
                "FAIL",
                "HARD_STOP",
                "<rows>",
                None,
                "The holdings table contains no positions.",
                "Provide the dated holdings table required by the selected input mode.",
            )
        )
    for index, row in enumerate(bundle.holdings, start=2):
        checks.extend(
            schema_checks(
                "holdings",
                row,
                "holding_row.schema.json",
                row_number=index,
            )
        )
    if not bundle.opportunities:
        checks.append(
            PrivateInputCheck(
                "G4I-opportunities-empty",
                "opportunities",
                "FAIL",
                "HARD_STOP",
                "<rows>",
                None,
                "The opportunity-set table contains no alternatives.",
                "Provide a dated local opportunity set; missing opportunity cost is NOT_EVALUATED.",
            )
        )
    for index, row in enumerate(bundle.opportunities, start=2):
        checks.extend(
            schema_checks(
                "opportunities",
                row,
                "opportunity_row.schema.json",
                row_number=index,
            )
        )


def _classification_and_date_checks(
    bundle: PrivateInputBundle,
    checks: list[PrivateInputCheck],
) -> None:
    classification = bundle.manifest.get("data_classification")
    documents: list[tuple[str, Any]] = [
        ("policy", bundle.policy),
        ("constraints", bundle.constraint_inputs),
        ("approval", bundle.approval),
        ("freshness", bundle.freshness_attestation),
    ]
    documents.extend(("exposures", row) for row in bundle.exposures)
    documents.extend(("holdings", row) for row in bundle.holdings)
    documents.extend(("opportunities", row) for row in bundle.opportunities)
    mismatches = [
        document
        for document, payload in documents
        if not isinstance(payload, dict) or payload.get("data_classification") != classification
    ]
    add_check(
        checks,
        check_id="G4I-classification-consistency",
        document="bundle",
        passed=classification in SUPPORTED_CLASSIFICATIONS and not mismatches,
        field="data_classification",
        message_pass="Data classification is consistent across all private-input documents.",
        message_fail="Data classification is missing, unsupported, or inconsistent across documents.",
        remediation="Use one classification across the manifest and every referenced document.",
    )

    as_of_date = bundle.manifest.get("as_of_date")
    date_mismatches = [
        document
        for document, payload in documents
        if not isinstance(payload, dict) or payload.get("as_of_date") != as_of_date
    ]
    add_check(
        checks,
        check_id="G4I-as-of-date-consistency",
        document="bundle",
        passed=parse_date(as_of_date) is not None and not date_mismatches,
        field="as_of_date",
        message_pass="All portfolio inputs use the manifest as-of date.",
        message_fail="One or more private inputs use a different or invalid as-of date.",
        remediation="Align policy, exposure, holdings, opportunity set, freshness attestation, and manifest dates.",
    )


def _policy_checks(bundle: PrivateInputBundle, checks: list[PrivateInputCheck]) -> None:
    policy = bundle.policy
    manifest = bundle.manifest
    as_of_date = parse_date(manifest.get("as_of_date"))
    expiration = parse_date(policy.get("expiration_review_date"))
    reviewed_at = parse_datetime(policy.get("reviewed_at"))
    chronology_valid = (
        as_of_date is not None
        and expiration is not None
        and expiration >= as_of_date
        and reviewed_at is not None
        and reviewed_at.date() <= as_of_date
    )
    add_check(
        checks,
        check_id="G4I-policy-chronology",
        document="policy",
        passed=chronology_valid,
        field="expiration_review_date,reviewed_at",
        message_pass="Policy review and expiration dates are chronologically valid.",
        message_fail="Policy review or expiration chronology is invalid.",
        remediation="Use a review timestamp no later than the as-of date and a non-expired review date.",
    )
    add_check(
        checks,
        check_id="G4I-policy-base-currency",
        document="policy",
        passed=policy.get("base_currency") == manifest.get("base_currency"),
        field="base_currency",
        message_pass="Policy and manifest base currencies match.",
        message_fail="Policy and manifest base currencies do not match.",
        remediation="Convert holdings to one base currency and use it consistently.",
    )
    hedge_instruments = policy.get("permitted_hedge_instruments")
    maximum_hedge_ratio = parse_ratio(policy.get("maximum_hedge_ratio"))
    hedge_policy_valid = (
        isinstance(hedge_instruments, list)
        and maximum_hedge_ratio is not None
        and (maximum_hedge_ratio == 0 or bool(hedge_instruments))
    )
    add_check(
        checks,
        check_id="G4I-hedge-policy-completeness",
        document="policy",
        passed=hedge_policy_valid,
        field="permitted_hedge_instruments,maximum_hedge_ratio",
        message_pass="Hedge permissions and maximum ratio are internally consistent.",
        message_fail="A positive hedge ratio requires at least one explicitly permitted instrument.",
        remediation="List permitted hedge instruments or set the maximum hedge ratio to zero.",
    )


def _exposure_checks(
    bundle: PrivateInputBundle,
    checks: list[PrivateInputCheck],
) -> None:
    if not bundle.exposures:
        if bundle.manifest.get("input_mode") == "FULL_HOLDINGS":
            checks.append(
                PrivateInputCheck(
                    check_id="G4I-exposures-optional-not-supplied",
                    document="exposures",
                    status="PASS",
                    issue_class="INFO",
                    field="<rows>",
                    row_number=None,
                    message="No independent exposure table was supplied; full holdings remain the authoritative source.",
                    remediation="None.",
                )
            )
        return

    mode = bundle.manifest.get("input_mode")
    base_currency = bundle.manifest.get("base_currency")
    nav = parse_number(bundle.manifest.get("portfolio_nav"))
    tolerance = parse_number(bundle.manifest.get("weight_reconciliation_tolerance"))
    exposure_ids = [clean_text(row.get("exposure_id")) for row in bundle.exposures]
    exposure_pairs = [
        (clean_text(row.get("exposure_type")), clean_text(row.get("exposure_key")))
        for row in bundle.exposures
    ]
    add_check(
        checks,
        check_id="G4I-exposure-id-uniqueness",
        document="exposures",
        passed=(
            bool(exposure_ids)
            and None not in exposure_ids
            and len(exposure_ids) == len(set(exposure_ids))
        ),
        field="exposure_id",
        message_pass="Exposure row IDs are present and unique.",
        message_fail="Exposure row IDs are missing or duplicated.",
        remediation="Assign one stable unique exposure_id to each row.",
    )
    add_check(
        checks,
        check_id="G4I-exposure-bucket-uniqueness",
        document="exposures",
        passed=(
            None not in {part for pair in exposure_pairs for part in pair}
            and len(exposure_pairs) == len(set(exposure_pairs))
        ),
        field="exposure_type,exposure_key",
        message_pass="Exposure type and key pairs are unique.",
        message_fail="One or more exposure buckets are duplicated.",
        remediation="Consolidate duplicate rows before validation.",
    )
    add_check(
        checks,
        check_id="G4I-exposure-base-currency",
        document="exposures",
        passed=all(
            row.get("base_currency") == base_currency for row in bundle.exposures
        ),
        field="base_currency",
        message_pass="Every exposure uses the manifest base currency.",
        message_fail="One or more exposures use a different base currency.",
        remediation="Convert market values to the manifest base currency.",
    )
    required_aggregate_types = {"GROSS", "NET"}
    supplied_types = {row.get("exposure_type") for row in bundle.exposures}
    add_check(
        checks,
        check_id="G4I-exposure-gross-net-coverage",
        document="exposures",
        passed=(
            mode == "FULL_HOLDINGS"
            or required_aggregate_types.issubset(supplied_types)
        ),
        field="exposure_type",
        message_pass="Required gross and net aggregate exposure rows are present.",
        message_fail="Exposure-only and aggregated modes require GROSS and NET rows.",
        remediation="Add reviewed GROSS and NET exposure rows.",
    )

    gross_rows = [
        row for row in bundle.exposures if row.get("exposure_type") == "GROSS"
    ]
    net_rows = [
        row for row in bundle.exposures if row.get("exposure_type") == "NET"
    ]
    gross = (
        parse_ratio(gross_rows[0].get("exposure_weight"))
        if len(gross_rows) == 1
        else None
    )
    net = (
        parse_ratio(net_rows[0].get("exposure_weight"))
        if len(net_rows) == 1
        else None
    )
    aggregate_relationship_valid = (
        mode == "FULL_HOLDINGS" and not gross_rows and not net_rows
    ) or (
        gross is not None
        and net is not None
        and gross >= 0
        and abs(net) <= gross
    )
    add_check(
        checks,
        check_id="G4I-exposure-gross-net-relationship",
        document="exposures",
        passed=aggregate_relationship_valid,
        field="exposure_weight",
        message_pass="Gross and net exposure relationships are internally consistent.",
        message_fail="Gross exposure is negative, duplicated, missing, or below absolute net exposure.",
        remediation="Correct the reviewed GROSS and NET rows without summing overlapping exposure dimensions.",
    )

    as_of_date = parse_date(bundle.manifest.get("as_of_date"))
    for index, row in enumerate(bundle.exposures, start=2):
        reviewed_at = parse_datetime(row.get("reviewed_at"))
        add_check(
            checks,
            check_id=f"G4I-exposure-review-chronology-{index}",
            document="exposures",
            passed=(
                reviewed_at is not None
                and as_of_date is not None
                and reviewed_at.date() <= as_of_date
            ),
            field="reviewed_at",
            row_number=index,
            message_pass="Exposure review timestamp is valid for the as-of date.",
            message_fail="Exposure review timestamp is missing or later than the as-of date.",
            remediation="Use a completed local review timestamp no later than the as-of date.",
        )
        if mode == "AGGREGATED_PORTFOLIO":
            weight = parse_ratio(row.get("exposure_weight"))
            market_value = parse_number(row.get("market_value_base_currency"))
            add_check(
                checks,
                check_id=f"G4I-exposure-nav-reconciliation-{index}",
                document="exposures",
                passed=(
                    nav is not None
                    and nav > 0
                    and tolerance is not None
                    and weight is not None
                    and market_value is not None
                    and abs(market_value / nav - weight) <= tolerance
                ),
                field="exposure_weight,market_value_base_currency",
                row_number=index,
                message_pass="Exposure weight reconciles to market value and NAV.",
                message_fail="Exposure weight does not reconcile to market value and NAV.",
                remediation="Correct the aggregate exposure weight or base-currency market value.",
            )
        source_basis = row.get("source_basis")
        add_check(
            checks,
            check_id=f"G4I-exposure-source-basis-{index}",
            document="exposures",
            passed=not (
                mode == "EXPOSURE_ONLY"
                and source_basis == "FULL_HOLDINGS_DERIVED"
            ),
            field="source_basis",
            row_number=index,
            message_pass="Exposure source basis is compatible with the selected input mode.",
            message_fail="Exposure-only mode cannot claim derivation from holdings that were not supplied.",
            remediation="Use the actual reviewed source basis or select the correct input mode.",
        )


def _holding_checks(bundle: PrivateInputBundle, checks: list[PrivateInputCheck]) -> None:
    manifest = bundle.manifest
    mode = manifest.get("input_mode")
    if mode == "EXPOSURE_ONLY":
        checks.append(
            PrivateInputCheck(
                check_id="G4I-holdings-prohibited-not-supplied",
                document="holdings",
                status="PASS",
                issue_class="INFO",
                field="<rows>",
                row_number=None,
                message="No holdings table was loaded in exposure-only mode.",
                remediation="None.",
            )
        )
        return
    tolerance = parse_number(manifest.get("weight_reconciliation_tolerance"))
    nav = parse_number(manifest.get("portfolio_nav"))
    base_currency = manifest.get("base_currency")
    position_ids = [clean_text(row.get("position_id")) for row in bundle.holdings]
    duplicate_ids = len(position_ids) != len(set(position_ids))
    add_check(
        checks,
        check_id="G4I-holdings-granularity",
        document="holdings",
        passed=all(
            row.get("position_granularity")
            == (
                "AGGREGATED_ISSUER"
                if mode == "AGGREGATED_PORTFOLIO"
                else "SECURITY_LEVEL"
            )
            for row in bundle.holdings
        ),
        field="position_granularity",
        message_pass="Every holding row matches the selected portfolio granularity.",
        message_fail="One or more holding rows use a granularity incompatible with the input mode.",
        remediation="Use AGGREGATED_ISSUER rows only in aggregated mode and SECURITY_LEVEL rows only in full-holdings mode.",
    )
    add_check(
        checks,
        check_id="G4I-holdings-position-id-uniqueness",
        document="holdings",
        passed=bool(position_ids) and not duplicate_ids and None not in position_ids,
        field="position_id",
        message_pass="Holding position IDs are present and unique.",
        message_fail="Holding position IDs are missing or duplicated.",
        remediation="Assign one stable unique position_id to every row.",
    )
    currency_consistent = all(row.get("base_currency") == base_currency for row in bundle.holdings)
    add_check(
        checks,
        check_id="G4I-holdings-base-currency",
        document="holdings",
        passed=currency_consistent,
        field="base_currency",
        message_pass="Every holding uses the manifest base currency.",
        message_fail="One or more holdings use a different base currency.",
        remediation="Convert every market value to the manifest base currency before validation.",
    )

    weights = [parse_ratio(row.get("position_weight")) for row in bundle.holdings]
    weight_total_valid = (
        tolerance is not None
        and all(value is not None for value in weights)
        and abs(sum(value for value in weights if value is not None) - 1.0) <= tolerance
    )
    add_check(
        checks,
        check_id="G4I-holdings-weight-reconciliation",
        document="holdings",
        passed=weight_total_valid,
        field="position_weight",
        message_pass="Portfolio weights reconcile to 100% within the explicit tolerance.",
        message_fail="Portfolio weights do not reconcile within the explicit tolerance.",
        remediation="Include all aggregate or security-level positions and cash, then reconcile signed weights to portfolio NAV.",
    )

    for index, row in enumerate(bundle.holdings, start=2):
        weight = parse_ratio(row.get("position_weight"))
        market_value = parse_number(row.get("market_value_base_currency"))
        row_reconciles = (
            nav is not None
            and nav > 0
            and tolerance is not None
            and weight is not None
            and market_value is not None
            and abs(market_value / nav - weight) <= tolerance
        )
        add_check(
            checks,
            check_id=f"G4I-holdings-nav-reconciliation-{index}",
            document="holdings",
            passed=row_reconciles,
            field="position_weight,market_value_base_currency",
            row_number=index,
            message_pass="The row weight reconciles to base-currency market value and NAV.",
            message_fail="The row weight does not reconcile to base-currency market value and NAV.",
            remediation="Correct the row weight or base-currency market value.",
        )
        side = row.get("position_side")
        sign_valid = (
            weight is not None
            and (
                (side in {"LONG", "CASH"} and weight >= 0)
                or (side == "SHORT" and weight <= 0)
                or side == "HEDGE"
            )
        )
        add_check(
            checks,
            check_id=f"G4I-holdings-side-sign-{index}",
            document="holdings",
            passed=sign_valid,
            field="position_side,position_weight",
            row_number=index,
            message_pass="Position side and signed weight are consistent.",
            message_fail="Position side and signed weight are inconsistent.",
            remediation="Use nonnegative LONG/CASH weights and nonpositive SHORT weights.",
        )
        hedge_ratio = parse_ratio(row.get("hedge_ratio"))
        hedge_reference_valid = (
            (mode == "AGGREGATED_PORTFOLIO" and hedge_ratio is None)
            or (
                hedge_ratio is not None
                and (
                    hedge_ratio == 0
                    or clean_text(row.get("existing_hedge_identifier")) is not None
                    or side == "HEDGE"
                )
            )
        )
        add_check(
            checks,
            check_id=f"G4I-holdings-hedge-reference-{index}",
            document="holdings",
            passed=hedge_reference_valid,
            field="existing_hedge_identifier,hedge_ratio",
            row_number=index,
            message_pass="Hedge ratio and hedge reference are consistent.",
            message_fail="A nonzero hedge ratio lacks a hedge identifier.",
            remediation="Provide the existing hedge identifier or set hedge_ratio to zero.",
        )


def _opportunity_checks(bundle: PrivateInputBundle, checks: list[PrivateInputCheck]) -> None:
    opportunity_ids = [clean_text(row.get("opportunity_id")) for row in bundle.opportunities]
    unique_ids = len(opportunity_ids) == len(set(opportunity_ids)) and None not in opportunity_ids
    add_check(
        checks,
        check_id="G4I-opportunity-id-uniqueness",
        document="opportunities",
        passed=bool(opportunity_ids) and unique_ids,
        field="opportunity_id",
        message_pass="Opportunity IDs are present and unique.",
        message_fail="Opportunity IDs are missing or duplicated.",
        remediation="Assign one stable unique opportunity_id to every row.",
    )
    required_count = parse_integer(
        bundle.policy.get("opportunity_cost_requirement", {}).get(
            "minimum_comparable_opportunities"
        )
    )
    validated_active_count = sum(
        1
        for row in bundle.opportunities
        if row.get("opportunity_status") == "ACTIVE"
        and row.get("return_status") == "VALIDATED"
        and parse_ratio(row.get("expected_return")) is not None
        and parse_ratio(row.get("bear_case_downside")) is not None
    )
    comparable_set_ready = (
        required_count is not None and validated_active_count >= required_count
    )
    add_check(
        checks,
        check_id="G4I-opportunity-cost-readiness",
        document="opportunities",
        passed=comparable_set_ready,
        field="return_status,opportunity_status",
        message_pass="The opportunity set contains the required validated active alternatives.",
        message_fail="The opportunity set lacks the required validated active alternatives.",
        remediation="Add dated alternatives with validated return and downside inputs; otherwise opportunity cost is NOT_EVALUATED.",
    )


def _constraint_input_checks(
    bundle: PrivateInputBundle,
    checks: list[PrivateInputCheck],
) -> None:
    payload = bundle.constraint_inputs
    candidate = payload.get("candidate", {})
    portfolio_state = payload.get("portfolio_state", {})
    binding = payload.get("gate3_binding", {})
    freshness = bundle.freshness_attestation
    manifest_date = parse_date(bundle.manifest.get("as_of_date"))

    add_check(
        checks,
        check_id="G4I-constraint-gate3-binding",
        document="constraints",
        passed=(
            isinstance(binding, dict)
            and binding.get("report_id") == freshness.get("gate3_report_id")
            and binding.get("contract_hash") == freshness.get("gate3_contract_hash")
        ),
        field="gate3_binding",
        message_pass="Constraint inputs are bound to the same Gate 3 identity as the freshness attestation.",
        message_fail="Constraint inputs and the freshness attestation reference different Gate 3 objects.",
        remediation="Repeat the private review and bind both documents to the exact consumed Gate 3 report ID and hash.",
    )

    expected = candidate.get("expected_return", {}) if isinstance(candidate, dict) else {}
    downside = candidate.get("downside_return", {}) if isinstance(candidate, dict) else {}
    both_validated = (
        isinstance(expected, dict)
        and isinstance(downside, dict)
        and expected.get("status") == "VALIDATED"
        and downside.get("status") == "VALIDATED"
    )
    horizons_match = (
        not both_validated
        or (
            expected.get("as_of_date") == downside.get("as_of_date")
            and expected.get("target_date") == downside.get("target_date")
            and expected.get("holding_period_days")
            == downside.get("holding_period_days")
        )
    )
    add_check(
        checks,
        check_id="G4I-constraint-return-horizon-alignment",
        document="constraints",
        passed=horizons_match,
        field="candidate.expected_return,candidate.downside_return",
        message_pass="Validated expected-return and downside inputs use one dated holding-period basis.",
        message_fail="Expected-return and downside inputs use different dates or holding periods.",
        remediation="Rebuild both return inputs on the same as-of date, target date, and holding period.",
    )

    for label, return_input in (
        ("expected-return", expected),
        ("downside-return", downside),
    ):
        if not isinstance(return_input, dict) or return_input.get("status") != "VALIDATED":
            checks.append(
                PrivateInputCheck(
                    check_id=f"G4I-constraint-{label}-explicit-status",
                    document="constraints",
                    status="PASS",
                    issue_class="INFO",
                    field=f"candidate.{label.replace('-', '_')}.status",
                    row_number=None,
                    message="The return input is explicitly marked provisional or missing; S13 must keep dependent constraints incomplete.",
                    remediation="None.",
                )
            )
            continue
        as_of = parse_date(return_input.get("as_of_date"))
        target = parse_date(return_input.get("target_date"))
        holding_days = parse_integer(return_input.get("holding_period_days"))
        reviewed_at = parse_datetime(return_input.get("reviewed_at"))
        calculated_days = (target - as_of).days if as_of and target else None
        chronology_valid = (
            manifest_date is not None
            and as_of is not None
            and target is not None
            and as_of <= manifest_date
            and target > as_of
            and holding_days is not None
            and calculated_days == holding_days
            and reviewed_at is not None
            and reviewed_at.date() <= manifest_date
        )
        add_check(
            checks,
            check_id=f"G4I-constraint-{label}-chronology",
            document="constraints",
            passed=chronology_valid,
            field=f"candidate.{label.replace('-', '_')}.as_of_date,target_date,holding_period_days,reviewed_at",
            message_pass="The validated return input has reproducible dates, holding period, and review chronology.",
            message_fail="The return dates, holding period, or review chronology do not reproduce.",
            remediation="Use target date minus as-of date as holding_period_days and complete review no later than the manifest date.",
        )
        add_check(
            checks,
            check_id=f"G4I-constraint-{label}-contract-binding",
            document="constraints",
            passed=return_input.get("source_contract_hash") == binding.get("contract_hash"),
            field=f"candidate.{label.replace('-', '_')}.source_contract_hash",
            message_pass="The validated return input is bound to the selected Gate 3 contract.",
            message_fail="The return input references a different or missing Gate 3 contract hash.",
            remediation="Revalidate the return input against the exact Gate 3 contract consumed by S13.",
        )

    liquidity = candidate.get("liquidity", {}) if isinstance(candidate, dict) else {}
    if isinstance(liquidity, dict) and liquidity.get("status") == "VALIDATED":
        liquidity_date = parse_date(liquidity.get("source_as_of_date"))
        liquidity_reviewed_at = parse_datetime(liquidity.get("reviewed_at"))
        maximum_advt_age = parse_integer(
            bundle.policy.get("liquidity_requirement", {}).get(
                "maximum_advt_age_days"
            )
        )
        advt_age = (
            (manifest_date - liquidity_date).days
            if manifest_date is not None and liquidity_date is not None
            else None
        )
        liquidity_valid = (
            manifest_date is not None
            and liquidity_date is not None
            and advt_age is not None
            and maximum_advt_age is not None
            and 0 <= advt_age <= maximum_advt_age
            and liquidity_reviewed_at is not None
            and liquidity_reviewed_at.date() <= manifest_date
            and liquidity.get("base_currency") == bundle.manifest.get("base_currency")
        )
        add_check(
            checks,
            check_id="G4I-constraint-liquidity-chronology-currency",
            document="constraints",
            passed=liquidity_valid,
            field="candidate.liquidity",
            message_pass="Candidate liquidity uses current, reviewed ADVT in the portfolio base currency.",
            message_fail="Candidate liquidity is stale or has an invalid date, reviewer chronology, or currency.",
            remediation="Refresh ADVT within the policy age limit, convert it to the manifest base currency, and complete local review.",
        )

    reviewed_state_names = (
        "current_risk_budget_usage",
        "current_liquid_portfolio_weight",
    )
    for state_name in reviewed_state_names:
        state = (
            portfolio_state.get(state_name, {})
            if isinstance(portfolio_state, dict)
            else {}
        )
        if not isinstance(state, dict) or state.get("status") != "VALIDATED":
            continue
        reviewed_at = parse_datetime(state.get("reviewed_at"))
        add_check(
            checks,
            check_id=f"G4I-constraint-{state_name}-review",
            document="constraints",
            passed=(
                manifest_date is not None
                and reviewed_at is not None
                and reviewed_at.date() <= manifest_date
            ),
            field=f"portfolio_state.{state_name}.reviewed_at",
            message_pass="The portfolio-state input has valid reviewer chronology.",
            message_fail="The portfolio-state reviewer timestamp is missing or later than the manifest date.",
            remediation="Complete and date the local portfolio-state review no later than the manifest date.",
        )

    risk_state = (
        portfolio_state.get("current_risk_budget_usage", {})
        if isinstance(portfolio_state, dict)
        else {}
    )
    add_check(
        checks,
        check_id="G4I-constraint-risk-method",
        document="constraints",
        passed=(
            not isinstance(risk_state, dict)
            or risk_state.get("status") != "VALIDATED"
            or risk_state.get("methodology") == bundle.policy.get("risk_budget_method")
        ),
        field="portfolio_state.current_risk_budget_usage.methodology",
        message_pass="Current risk-budget usage and policy use the same method.",
        message_fail="Current risk-budget usage and policy use different methods.",
        remediation="Recalculate current usage using the policy-defined risk-budget method.",
    )

    hedge = candidate.get("proposed_hedge", {}) if isinstance(candidate, dict) else {}
    if isinstance(hedge, dict) and hedge.get("status") == "PROPOSED":
        hedge_reviewed_at = parse_datetime(hedge.get("reviewed_at"))
        add_check(
            checks,
            check_id="G4I-constraint-hedge-review",
            document="constraints",
            passed=(
                manifest_date is not None
                and hedge_reviewed_at is not None
                and hedge_reviewed_at.date() <= manifest_date
            ),
            field="candidate.proposed_hedge.reviewed_at",
            message_pass="The proposed hedge has valid reviewer chronology.",
            message_fail="The proposed hedge review is missing or later than the manifest date.",
            remediation="Complete the local hedge review no later than the manifest date.",
        )


def _approval_checks(
    bundle: PrivateInputBundle,
    checks: list[PrivateInputCheck],
    *,
    system_assessment_ready: bool,
) -> None:
    approval = bundle.approval
    manifest_as_of = parse_date(bundle.manifest.get("as_of_date"))
    approval_reviewed_at = parse_datetime(approval.get("reviewed_at"))
    add_check(
        checks,
        check_id="G4I-approval-review-chronology",
        document="approval",
        passed=(
            manifest_as_of is not None
            and approval_reviewed_at is not None
            and approval_reviewed_at.date() <= manifest_as_of
        ),
        field="reviewed_at",
        message_pass="Approval configuration review date is valid for the manifest date.",
        message_fail="Approval configuration review date is invalid for the manifest date.",
        remediation="Use a review timestamp no later than the workspace as-of date.",
    )
    allowed = set(approval.get("allowed_decisions", []))
    add_check(
        checks,
        check_id="G4I-approval-decision-set",
        document="approval",
        passed=allowed == ALLOWED_DECISIONS,
        field="allowed_decisions",
        message_pass="The complete controlled Partner decision set is configured.",
        message_fail="The configured Partner decision set is incomplete or unsupported.",
        remediation="Use exactly PENDING, APPROVED, APPROVED_WITH_MODIFICATION, REJECTED, and DEFERRED.",
    )
    decision = approval.get("partner_decision", {})
    status = decision.get("status") if isinstance(decision, dict) else None
    approved_by = clean_text(decision.get("approved_by")) if isinstance(decision, dict) else None
    approved_at = parse_datetime(decision.get("approved_at")) if isinstance(decision, dict) else None
    rationale = clean_text(decision.get("decision_rationale")) if isinstance(decision, dict) else None
    minimum = parse_ratio(decision.get("approved_position_min")) if isinstance(decision, dict) else None
    maximum = parse_ratio(decision.get("approved_position_max")) if isinstance(decision, dict) else None
    if status == "PENDING":
        decision_valid = all(value is None for value in (approved_by, approved_at, rationale, minimum, maximum))
    elif status in {"APPROVED", "APPROVED_WITH_MODIFICATION"}:
        decision_valid = (
            approved_by is not None
            and approved_at is not None
            and rationale is not None
            and minimum is not None
            and maximum is not None
            and 0 <= minimum <= maximum <= 1
            and manifest_as_of is not None
            and approved_at.date() <= manifest_as_of
        )
    elif status in {"REJECTED", "DEFERRED"}:
        decision_valid = (
            approved_by is not None
            and approved_at is not None
            and rationale is not None
            and minimum is None
            and maximum is None
            and manifest_as_of is not None
            and approved_at.date() <= manifest_as_of
        )
    else:
        decision_valid = False
    add_check(
        checks,
        check_id="G4I-partner-decision-completeness",
        document="approval",
        passed=decision_valid,
        field="partner_decision",
        message_pass="Partner decision fields are complete for the selected status.",
        message_fail="Partner decision fields are inconsistent with the selected status.",
        remediation="Keep all approval details null while pending; otherwise provide the required named approval record.",
    )
    add_check(
        checks,
        check_id="G4I-preassessment-decision-boundary",
        document="approval",
        passed=system_assessment_ready or status == "PENDING",
        field="partner_decision.status",
        message_pass="Partner decision status is valid for the current workflow stage.",
        message_fail="A Partner decision cannot be approved before the system assessment is ready.",
        remediation="Keep partner_decision.status as PENDING until the Gate 4 system assessment is complete.",
    )
    add_check(
        checks,
        check_id="G4I-no-automatic-trade",
        document="approval",
        passed=approval.get("automatic_trade_execution") is False,
        field="automatic_trade_execution",
        message_pass="Automatic trade execution is disabled.",
        message_fail="Automatic trade execution must remain disabled.",
        remediation="Set automatic_trade_execution to false.",
    )
    add_check(
        checks,
        check_id="G4I-external-transmission-denied",
        document="approval",
        passed=approval.get("external_transmission_policy") == "DENY",
        field="external_transmission_policy",
        message_pass="External transmission is denied by default.",
        message_fail="External transmission must be DENY for the local Gate 4 workflow.",
        remediation="Set external_transmission_policy to DENY.",
    )


def validate_private_input_bundle(
    bundle: PrivateInputBundle,
    *,
    existing_checks: list[PrivateInputCheck] | None = None,
    system_assessment_ready: bool = False,
) -> dict[str, Any]:
    checks = list(existing_checks or [])
    _validate_document_schemas(bundle, checks)
    _field_governance_checks(bundle, checks)
    _classification_and_date_checks(bundle, checks)
    _policy_checks(bundle, checks)
    _exposure_checks(bundle, checks)
    _holding_checks(bundle, checks)
    _opportunity_checks(bundle, checks)
    _constraint_input_checks(bundle, checks)
    _approval_checks(
        bundle,
        checks,
        system_assessment_ready=system_assessment_ready,
    )
    failed = [check for check in checks if check.status in {"FAIL", "MISSING", "BLOCKED"}]
    status = INPUT_STATUS_VALIDATED if not failed else INPUT_STATUS_REQUIRED
    classification = bundle.manifest.get("data_classification")
    input_mode = bundle.manifest.get("input_mode")
    return {
        "private_input_contract_version": PRIVATE_INPUT_CONTRACT_VERSION,
        "framework_status": FRAMEWORK_STATUS,
        "status": status,
        "validated_at": utc_now(),
        "data_classification": classification,
        "input_mode": input_mode,
        "mode_capabilities": MODE_CAPABILITIES.get(
            str(input_mode),
            {
                "aggregate_exposure_validation": "NOT_EVALUATED",
                "holdings_reconciliation": "NOT_EVALUATED",
                "security_level_liquidity": "NOT_EVALUATED",
            },
        ),
        "privacy_safe_diagnostic": True,
        "raw_values_included": False,
        "system_portfolio_assessment": "NOT_EVALUATED",
        "partner_approval_state": "PARTNER_APPROVAL_PENDING",
        "check_summary": {
            "total": len(checks),
            "passed": sum(check.status == "PASS" for check in checks),
            "failed": len(failed),
            "warnings": sum(check.status == "WARNING" for check in checks),
        },
        "blocking_check_ids": [check.check_id for check in failed],
        "checks": [check.to_dict() for check in checks],
        "next_action": (
            "Proceed to the local Gate 4 entry check."
            if status == INPUT_STATUS_VALIDATED
            else "Complete or correct the named private-input fields locally."
        ),
    }


def load_and_validate_private_inputs(
    manifest_path: Path,
) -> tuple[PrivateInputBundle | None, dict[str, Any]]:
    checks: list[PrivateInputCheck] = []
    try:
        manifest = read_mapping(manifest_path)
    except PrivateInputLoadError as exc:
        checks.append(
            PrivateInputCheck(
                "G4I-manifest-load",
                "manifest",
                "FAIL",
                "HARD_STOP",
                "<document>",
                None,
                str(exc),
                "Create the manifest from the public template and keep it local.",
            )
        )
        return None, _diagnostic_without_bundle(checks)
    checks.extend(
        schema_checks(
            "manifest",
            manifest,
            "private_workspace_manifest.schema.json",
        )
    )
    if any(check.status == "FAIL" for check in checks):
        return None, _diagnostic_without_bundle(checks)
    try:
        governance = load_field_governance_contract()
    except PrivateInputLoadError as exc:
        checks.append(
            PrivateInputCheck(
                "G4I-field-governance-load",
                "field-governance-contract",
                "FAIL",
                "HARD_STOP",
                "<contract>",
                None,
                str(exc),
                "Restore the shared public field-governance contract.",
            )
        )
        return None, _diagnostic_without_bundle(checks)
    _governance_definition_checks(governance, checks)
    if any(check.status == "FAIL" for check in checks):
        return None, _diagnostic_without_bundle(checks)
    bundle = _load_bundle_documents(
        manifest_path,
        manifest,
        governance,
        checks,
    )
    if bundle is None:
        return None, _diagnostic_without_bundle(checks)
    return bundle, validate_private_input_bundle(bundle, existing_checks=checks)


def _diagnostic_without_bundle(checks: list[PrivateInputCheck]) -> dict[str, Any]:
    failed = [check for check in checks if check.status in {"FAIL", "MISSING", "BLOCKED"}]
    return {
        "private_input_contract_version": PRIVATE_INPUT_CONTRACT_VERSION,
        "framework_status": FRAMEWORK_STATUS,
        "status": INPUT_STATUS_REQUIRED,
        "validated_at": utc_now(),
        "data_classification": "NOT_EVALUATED",
        "input_mode": "NOT_EVALUATED",
        "mode_capabilities": {
            "aggregate_exposure_validation": "NOT_EVALUATED",
            "holdings_reconciliation": "NOT_EVALUATED",
            "security_level_liquidity": "NOT_EVALUATED",
        },
        "privacy_safe_diagnostic": True,
        "raw_values_included": False,
        "system_portfolio_assessment": "NOT_EVALUATED",
        "partner_approval_state": "PARTNER_APPROVAL_PENDING",
        "check_summary": {
            "total": len(checks),
            "passed": sum(check.status == "PASS" for check in checks),
            "failed": len(failed),
            "warnings": sum(check.status == "WARNING" for check in checks),
        },
        "blocking_check_ids": [check.check_id for check in failed],
        "checks": [check.to_dict() for check in checks],
        "next_action": "Complete or correct the named private-input fields locally.",
    }
