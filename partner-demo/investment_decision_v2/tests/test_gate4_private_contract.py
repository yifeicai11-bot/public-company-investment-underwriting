#!/usr/bin/env python3
from __future__ import annotations

import csv
import json
import shutil
import sys
import tempfile
import unittest
from pathlib import Path

import yaml
from openpyxl import Workbook


TEST_DIR = Path(__file__).resolve().parent
SCRIPT_DIR = TEST_DIR.parent / "scripts"
GATE4_DIR = TEST_DIR.parent / "gate4"
SYNTHETIC_DIR = GATE4_DIR / "synthetic_examples"
TEMPLATE_DIR = GATE4_DIR / "templates"
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

from gate4_private_contract import (  # noqa: E402
    INPUT_STATUS_REQUIRED,
    INPUT_STATUS_VALIDATED,
    load_and_validate_private_inputs,
    read_mapping,
)


class Gate4PrivateContractTests(unittest.TestCase):
    def copy_synthetic_workspace(self, destination: Path) -> Path:
        workspace = destination / "synthetic_workspace"
        shutil.copytree(SYNTHETIC_DIR, workspace)
        return workspace / "synthetic_gate4_manifest.json"

    def test_synthetic_bundle_validates_without_raw_values_in_diagnostic(self) -> None:
        bundle, diagnostic = load_and_validate_private_inputs(
            SYNTHETIC_DIR / "synthetic_gate4_manifest.json"
        )
        serialized = json.dumps(diagnostic)

        self.assertIsNotNone(bundle)
        self.assertEqual(diagnostic["status"], INPUT_STATUS_VALIDATED)
        self.assertEqual(diagnostic["check_summary"]["failed"], 0)
        self.assertTrue(diagnostic["privacy_safe_diagnostic"])
        self.assertFalse(diagnostic["raw_values_included"])
        for private_value in (
            "Synthetic Alpha Corp",
            "SYNTH-ALPHA",
            "10000000",
            "target_return",
        ):
            self.assertNotIn(private_value, serialized)

    def test_missing_policy_field_returns_private_inputs_required(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            manifest_path = self.copy_synthetic_workspace(Path(tmp))
            policy_path = manifest_path.parent / "synthetic_portfolio_policy.yaml"
            policy = yaml.safe_load(policy_path.read_text(encoding="utf-8"))
            policy.pop("target_return")
            policy_path.write_text(yaml.safe_dump(policy, sort_keys=False), encoding="utf-8")

            _, diagnostic = load_and_validate_private_inputs(manifest_path)

        self.assertEqual(diagnostic["status"], INPUT_STATUS_REQUIRED)
        self.assertGreater(diagnostic["check_summary"]["failed"], 0)
        self.assertTrue(
            any(
                check["document"] == "policy" and "target_return" in check["field"]
                for check in diagnostic["checks"]
            )
        )

    def test_holdings_must_reconcile_to_nav_and_full_weight(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            manifest_path = self.copy_synthetic_workspace(Path(tmp))
            holdings_path = manifest_path.parent / "synthetic_current_holdings.csv"
            with holdings_path.open(newline="", encoding="utf-8") as handle:
                rows = list(csv.DictReader(handle))
                fieldnames = list(rows[0])
            rows[0]["position_weight"] = "0.31"
            with holdings_path.open("w", newline="", encoding="utf-8") as handle:
                writer = csv.DictWriter(handle, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(rows)

            _, diagnostic = load_and_validate_private_inputs(manifest_path)

        self.assertEqual(diagnostic["status"], INPUT_STATUS_REQUIRED)
        self.assertIn(
            "G4I-holdings-weight-reconciliation",
            diagnostic["blocking_check_ids"],
        )
        self.assertIn(
            "G4I-holdings-nav-reconciliation-2",
            diagnostic["blocking_check_ids"],
        )

    def test_opportunity_cost_is_not_evaluated_without_validated_alternatives(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            manifest_path = self.copy_synthetic_workspace(Path(tmp))
            opportunity_path = manifest_path.parent / "synthetic_opportunity_set.csv"
            with opportunity_path.open(newline="", encoding="utf-8") as handle:
                rows = list(csv.DictReader(handle))
                fieldnames = list(rows[0])
            for row in rows:
                row["return_status"] = "PROVISIONAL"
            with opportunity_path.open("w", newline="", encoding="utf-8") as handle:
                writer = csv.DictWriter(handle, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(rows)

            _, diagnostic = load_and_validate_private_inputs(manifest_path)

        self.assertEqual(diagnostic["status"], INPUT_STATUS_REQUIRED)
        self.assertIn(
            "G4I-opportunity-cost-readiness",
            diagnostic["blocking_check_ids"],
        )
        self.assertEqual(diagnostic["system_portfolio_assessment"], "NOT_EVALUATED")

    def test_pending_approval_cannot_contain_position_decision(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            manifest_path = self.copy_synthetic_workspace(Path(tmp))
            approval_path = manifest_path.parent / "synthetic_approval_config.yaml"
            approval = yaml.safe_load(approval_path.read_text(encoding="utf-8"))
            approval["partner_decision"]["approved_position_max"] = 0.02
            approval_path.write_text(
                yaml.safe_dump(approval, sort_keys=False),
                encoding="utf-8",
            )

            _, diagnostic = load_and_validate_private_inputs(manifest_path)

        self.assertEqual(diagnostic["status"], INPUT_STATUS_REQUIRED)
        self.assertIn(
            "G4I-partner-decision-completeness",
            diagnostic["blocking_check_ids"],
        )

    def test_xlsx_holdings_are_supported(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            manifest_path = self.copy_synthetic_workspace(Path(tmp))
            workspace = manifest_path.parent
            csv_path = workspace / "synthetic_current_holdings.csv"
            with csv_path.open(newline="", encoding="utf-8") as handle:
                rows = list(csv.reader(handle))
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "holdings"
            for row in rows:
                worksheet.append(row)
            xlsx_path = workspace / "synthetic_current_holdings.xlsx"
            workbook.save(xlsx_path)
            workbook.close()

            manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
            manifest["files"]["current_holdings"] = xlsx_path.name
            manifest_path.write_text(json.dumps(manifest, indent=2), encoding="utf-8")

            _, diagnostic = load_and_validate_private_inputs(manifest_path)

        self.assertEqual(diagnostic["status"], INPUT_STATUS_VALIDATED)
        self.assertEqual(diagnostic["check_summary"]["failed"], 0)

    def test_xlsx_formulas_are_rejected_without_evaluating_them(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            manifest_path = self.copy_synthetic_workspace(Path(tmp))
            workspace = manifest_path.parent
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.append(["data_classification", "as_of_date"])
            worksheet.append(["SYNTHETIC_PUBLIC_EXAMPLE", "=TODAY()"])
            xlsx_path = workspace / "synthetic_formula_holdings.xlsx"
            workbook.save(xlsx_path)
            workbook.close()

            manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
            manifest["files"]["current_holdings"] = xlsx_path.name
            manifest_path.write_text(json.dumps(manifest, indent=2), encoding="utf-8")

            bundle, diagnostic = load_and_validate_private_inputs(manifest_path)

        self.assertIsNone(bundle)
        self.assertEqual(diagnostic["status"], INPUT_STATUS_REQUIRED)
        self.assertIn("G4I-holdings-load", diagnostic["blocking_check_ids"])

    def test_xlsx_formula_header_is_rejected(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            manifest_path = self.copy_synthetic_workspace(Path(tmp))
            workspace = manifest_path.parent
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.append(["=LOWER(\"DATA_CLASSIFICATION\")", "as_of_date"])
            worksheet.append(["SYNTHETIC_PUBLIC_EXAMPLE", "2026-07-17"])
            xlsx_path = workspace / "synthetic_formula_header.xlsx"
            workbook.save(xlsx_path)
            workbook.close()

            manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
            manifest["files"]["current_holdings"] = xlsx_path.name
            manifest_path.write_text(json.dumps(manifest, indent=2), encoding="utf-8")

            bundle, diagnostic = load_and_validate_private_inputs(manifest_path)

        self.assertIsNone(bundle)
        self.assertEqual(diagnostic["status"], INPUT_STATUS_REQUIRED)
        self.assertIn("G4I-holdings-load", diagnostic["blocking_check_ids"])

    def test_duplicate_json_and_yaml_keys_are_rejected(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            json_path = root / "duplicate.json"
            yaml_path = root / "duplicate.yaml"
            json_path.write_text('{"schema_version": "1", "schema_version": "2"}', encoding="utf-8")
            yaml_path.write_text("schema_version: '1'\nschema_version: '2'\n", encoding="utf-8")

            with self.assertRaisesRegex(ValueError, "duplicate key"):
                read_mapping(json_path)
            with self.assertRaisesRegex(ValueError, "duplicate key"):
                read_mapping(yaml_path)

    def test_nonfinite_numeric_input_is_rejected(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            manifest_path = self.copy_synthetic_workspace(Path(tmp))
            holdings_path = manifest_path.parent / "synthetic_current_holdings.csv"
            with holdings_path.open(newline="", encoding="utf-8") as handle:
                rows = list(csv.DictReader(handle))
                fieldnames = list(rows[0])
            rows[0]["market_value_base_currency"] = "NaN"
            with holdings_path.open("w", newline="", encoding="utf-8") as handle:
                writer = csv.DictWriter(handle, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(rows)

            _, diagnostic = load_and_validate_private_inputs(manifest_path)

        self.assertEqual(diagnostic["status"], INPUT_STATUS_REQUIRED)
        self.assertTrue(
            any(
                check["document"] == "holdings"
                and "market_value_base_currency" in check["field"]
                for check in diagnostic["checks"]
                if check["status"] == "FAIL"
            )
        )

    def test_partner_approval_is_blocked_before_system_assessment(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            manifest_path = self.copy_synthetic_workspace(Path(tmp))
            approval_path = manifest_path.parent / "synthetic_approval_config.yaml"
            approval = yaml.safe_load(approval_path.read_text(encoding="utf-8"))
            approval["partner_decision"] = {
                "status": "APPROVED",
                "approved_by": "Synthetic Partner",
                "approved_at": "2026-07-17T08:00:00Z",
                "decision_rationale": "Synthetic approval record for boundary testing.",
                "approved_position_min": 0.01,
                "approved_position_max": 0.02,
            }
            approval_path.write_text(
                yaml.safe_dump(approval, sort_keys=False),
                encoding="utf-8",
            )

            _, diagnostic = load_and_validate_private_inputs(manifest_path)

        self.assertEqual(diagnostic["status"], INPUT_STATUS_REQUIRED)
        self.assertIn(
            "G4I-preassessment-decision-boundary",
            diagnostic["blocking_check_ids"],
        )

    def test_empty_public_template_does_not_invent_policy_defaults(self) -> None:
        _, diagnostic = load_and_validate_private_inputs(
            TEMPLATE_DIR / "gate4_private_workspace_manifest.template.json"
        )

        self.assertEqual(diagnostic["status"], INPUT_STATUS_REQUIRED)
        self.assertGreater(diagnostic["check_summary"]["failed"], 0)
        self.assertEqual(diagnostic["system_portfolio_assessment"], "NOT_EVALUATED")


if __name__ == "__main__":
    unittest.main()
